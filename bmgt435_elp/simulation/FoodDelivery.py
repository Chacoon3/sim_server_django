import openpyxl
import io
import numpy as np
import pandas as pd
import scipy.stats
from .Core import CaseBase, SimulationException, SimulationResult


class FoodDeliveryResult(SimulationResult):

    def detail_as_excel_stream(self):
        wb = openpyxl.Workbook(write_only=True)
        main_sheet = wb.create_sheet('main')
        main_sheet.append(self.aggregation_dataframe.columns.tolist())
        for row in self.aggregation_dataframe.values.tolist():
            main_sheet.append(row)
        
        detail_sheet = wb.create_sheet('detail')
        detail_sheet.append(self.iteration_dataframe.columns.tolist())
        for row in self.iteration_dataframe.values.tolist():
            detail_sheet.append(row)
    
        bytes_io = io.BytesIO()
        wb.save(filename=bytes_io)
        bytes_io.seek(0)
        return bytes_io
    

    def summary_as_dict(self):
        return self.aggregation_dataframe.loc[0,].to_dict()


class FoodDelivery(CaseBase):

    class DeliveryHub(object):
        """
        maintain the state of a delivery hub
        """

        def __init__(self, policy: tuple[int, int], initial_inventory, name) -> None:
            self.__s_small = policy[0]
            self.__s_big = policy[1]
            self.__inventory = initial_inventory
            self.__name = name

        def s_small(self): return self.__s_small

        def s_big(self): return self.__s_big

        def get_inventory(self): return self.__inventory

        def set_inventory(self, val):
            if val < 0:
                raise Exception("inventory cannot be negative")

            self.__inventory = val

        def add_inventory(self, val):
            self.set_inventory(self.__inventory + val)
            return

        def get_name(self):
            return self.__name

    # static members

    # original entities are ['10', '13', '43', '52', '67', '137'] from the dataset
    __center_names = [str(i) for i in range(1, 7)]
    __center_weekly_cost: int = 24000    # weekly fixed cost for a center
    __demand_mu_vector: list[int] = [2329, 2967, 2711, 2153, 1958, 2155]
    __demand_cov_matrix: np.ndarray = np.array([
        [113652.9946,	-37465.01062,	152.2425135,
            102.7690763,	32011.70125,	89.03852468],
        [-37465.01062,	137223.4483,	100715.5776,
            111.7531877,	-23449.91204,	103.1471614],
        [152.2425135,	100715.5776,	295682.0494,
            151.5108562,	134.5415465,	-75566.93011],
        [102.7690763,	111.7531877,	151.5108562,
            137073.65,	101.4975821,	94.87283555],
        [32011.70125,	-23449.91204,	134.5415465,
            101.4975821,	100183.0197,	91.92860568],
        [89.03852468,	103.1471614,	-75566.93011,
            94.87283555,	91.92860568,	120703.1537]
    ])
    __holding_cost: float = 10     # multiplier for holding cost
    __min_week_demand: int = 10
    __max_week_demand: int = 6000
    __num_weeks: int = 52
    __initial_inventory: int = 1000
    __max_weekly_restock: int = 7000
    __num_iterations: int = 1

    @staticmethod
    def __sim_center_demand():
        """
        returns discretized, truncated demand of all the six centers regardless of whether it is chosen, stored in a dict where keys are center ids
        """

        arr_demand = scipy.stats.multivariate_normal.rvs(
            mean=FoodDelivery.__demand_mu_vector, cov=FoodDelivery.__demand_cov_matrix, size=1)
        return {
            center_name: min(max(FoodDelivery.__min_week_demand, round(center_demand)),
                   FoodDelivery.__max_week_demand)
            for center_name, center_demand in zip(FoodDelivery.__center_names, arr_demand)
        }

    @staticmethod
    def __sim_checkout_price(size: int = 1):
        price = np.random.triangular(6.7, 29, 76.6, size)
        price = np.round(price, decimals=2)
        return price
    


    def __init__(
        self,
        centers: list[str] = [],
        policies: list[tuple[int, int]] | list[list[int]] = [],
        input_mapper: dict | None = None,
    ):

        super().__init__()
        self.__centers = centers
        self.__policies = policies
        self.__input_mapper = input_mapper
        self._assert_params()

    def _assert_params(self) -> None:
        config_valid = self.__num_iterations > 0 and self.__num_weeks > 0

        format_valid = len(self.__centers) == len(self.__policies) and len(self.__centers) > 0

        location_valid = all([l in FoodDelivery.__center_names for l in self.__centers])

        policy_valid = all([p[0] >= 0 and p[1] > p[0] for p in self.__policies])

        if self.__input_mapper is  not None:
            s_keys = set(self.__input_mapper.keys())
            s_values = set(self.__input_mapper.values())
            s_valid_values = set(FoodDelivery.__center_names)
            input_mapper_valid = s_keys == s_values and s_keys == s_valid_values
            if not input_mapper_valid:
                raise SimulationException("Simulation failed. The input mapper is not valid!")

        if not config_valid:
            raise SimulationException(
                "Simulation failed. The current case config is invalid. Please report this to the administrator.")
        elif not format_valid:
            raise SimulationException(
                "Invalid parameters. Please select at least one location and assign s-S policies to each location.")
        elif not location_valid:
            raise SimulationException(
                "Simulation failed. The input location arguments are not valid!")
        elif not policy_valid:
            raise SimulationException(
                "Simulation failed. s policy must be non-negative and S policy must be greater than s policy")

    perf_lower_bound = -800000
    perf_upper_bound = 1200000


    def score(self, obj) -> float:
        """
        returns the score of the simulation
        """
        avg_profit = obj['perf_metric']
        avg_profit = (avg_profit - FoodDelivery.perf_lower_bound) / \
            (FoodDelivery.perf_upper_bound - FoodDelivery.perf_lower_bound)
        avg_profit = 1 / (1 + np.exp(-avg_profit))
        avg_profit = round(avg_profit * 100, 3) 
        return avg_profit
    

    def meta_data(self):
        return {
            'num_iterations': self.__num_iterations,
            'num_weeks': self.__num_weeks,
        }

    def simulate(self):
        # instantialize centers
        centers = [
            FoodDelivery.DeliveryHub(policy=policy, initial_inventory=FoodDelivery.__initial_inventory, name=center) for \
            policy, center in zip(self.__policies, self.__centers)
        ]
        # define the output data:
        #       the output data describes one iteration of the simulation
        #       a nested dictionary is used to store the weekly state of each center and thereby yield the aggregation statistics
        history = {
            center.get_name(): {
                # 'cum_demand': 0,
                # 'cum_supply': 0,
                # 'cum_shortage_count': 0,
                # 'cum_shortage_amount': 0,
                # 'cum_revenue': 0,
                # 'cum_holding_cost': 0,

                # keys below are for detailed output
                'prior_inventory': [],
                'post_inventory': [],
                'demand': [],
                'supply': [],
                'shortage_count': [],
                'shortage_amount': [],
                'revenue': [],
                'holding_cost': []
            }
            for center in centers
        }

        output = {
            'perf_metric': float('-inf'),
            'total_revenue': 0,
            'total_shortage_count': 0,
            'total_shortage_amount': 0,
            'total_holding_cost': 0,
            'total_fixed_cost': 0,
            'history': history
        }

        # main logic

        # for every week
        for every_week in range(FoodDelivery.__num_weeks):

            # check inventory
            arr_purchase = [center.s_big() - center.get_inventory() if center.get_inventory()
                            <= center.s_small() else 0 for center in centers]   # has to be <= here otherwise a small s equal to 0 will not work as expected
            total_purchase = sum(arr_purchase)

            # check if inventory request exceeds the constraint
            if total_purchase > FoodDelivery.__max_weekly_restock:
                # if exceeds the constraint, adjust the original request array proportionally so that the sum equals the constraint
                purchase_ratio = np.array(arr_purchase) / total_purchase
                adjusted_purchase = purchase_ratio * FoodDelivery.__max_weekly_restock
                adjusted_purchase = np.round(
                    adjusted_purchase, decimals=0).astype(int)
                total_purchase = sum(adjusted_purchase)

                # in case the conversion from float to int leads the adjusted purchase to exceed the constraint
                # reduce the purchase amout of each center till the sum equals the constraint again
                if total_purchase > FoodDelivery.__max_weekly_restock:
                    diff = total_purchase - FoodDelivery.__max_weekly_restock
                    while diff > 0:
                        for i in range(len(adjusted_purchase)):
                            if (adjusted_purchase[i] > 0):
                                if adjusted_purchase[i] >= diff:
                                    adjusted_purchase[i] -= diff
                                    diff = 0
                                else:
                                    diff -= adjusted_purchase[i]
                                    adjusted_purchase[i] = 0
                            if diff == 0:
                                break
                arr_purchase = adjusted_purchase

            # update inventory
            for center, purchase in zip(centers, arr_purchase):
                center.add_inventory(purchase)

            # decide the weekly demand at each center
            #center_demand = np.array([self.__dict_center_demand[center.get_name()]() for center in centers])
            center_demand = FoodDelivery.__sim_center_demand()
            center_demand = [center_demand[center.get_name()]
                             for center in centers]

            # decide the actual number of orders that will be handled by each center
            center_supply = np.array([min(center.get_inventory(), demand)
                                     for center, demand in zip(centers, center_demand)])

            # for each center:
            #       decide the observations of checkout price
            #       calculate the weekly revenue
            #       apply shortage penalty
            #       record metrics of interest into output
            for i in range(len(centers)):

                center = centers[i]
                c_name = center.get_name()
                demand = center_demand[i]
                supply = center_supply[i]
                shortage_count = max(0, demand - supply)
                arr_order_price = FoodDelivery.__sim_checkout_price(size=demand)
                # orders covered
                covered_order_price = arr_order_price[:supply]
                # orders failed to cover
                lost_order_price = arr_order_price[supply:]
                order_revenue = round(sum(covered_order_price), 2)
                shortage_penalty = round(sum(lost_order_price), 2)

                # this records the inventory level after the weekly purchase
                prior_inv = center.get_inventory()
                center.add_inventory(-supply)
                # this records the inventory after supplying the demand in the week
                post_inv = center.get_inventory()
                holding_cost = post_inv * FoodDelivery.__holding_cost

                # history[c_name]['cum_demand'] += demand
                # history[c_name]['cum_supply'] += supply
                # history[c_name]['cum_shortage_count'] += shortage_count
                # history[c_name]['cum_shortage_amount'] = round(
                #     history[c_name]['cum_shortage_amount'] + shortage_penalty, 2)
                # history[c_name]['cum_revenue'] = round(
                #     history[c_name]['cum_revenue'] + order_revenue, 2)
                # history[c_name]['cum_holding_cost'] += holding_cost

                history[c_name]['prior_inventory'].append(prior_inv)
                history[c_name]['post_inventory'].append(post_inv)
                history[c_name]['demand'].append(demand)
                history[c_name]['supply'].append(supply)
                history[c_name]['shortage_count'].append(shortage_count)
                history[c_name]['shortage_amount'].append(shortage_penalty)
                history[c_name]['revenue'].append(order_revenue)
                history[c_name]['holding_cost'].append(holding_cost)

        # perform aggregation
        output['total_revenue'] = round(sum([
            sum(history[c.get_name()]['revenue']) for c in centers
        ]), 2)
        output['total_shortage_count'] = round(sum([
            sum(history[c.get_name()]['shortage_count']) for c in centers
        ]), 2)
        output['total_shortage_amount'] = round(sum([
            sum(history[c.get_name()]['shortage_amount']) for c in centers
        ]), 2)
        output['total_holding_cost'] = sum([
            sum(history[c.get_name()]['holding_cost']) for c in centers
        ])
        output['total_fixed_cost'] = len(
            centers) * self.__num_weeks * FoodDelivery.__center_weekly_cost
        output['perf_metric'] = round(
            output['total_revenue'] - output['total_shortage_amount'] -
            output['total_fixed_cost'] - output['total_holding_cost'],
            2
        )
        return output

    def run(self):
        if self.__input_mapper is not None:
            self.__centers = [self.__input_mapper[c] for c in self.__centers]
        
        res = self.simulate()
        score = self.score(res)
        history = res.pop('history')
        df_aggregated_statistics = pd.DataFrame(res, index=[0])
        arr_df_per_center_statistics = [
            pd.DataFrame(history[center_name]) for center_name in history.keys()
        ]
        for centerwise_df, c_name in zip(arr_df_per_center_statistics, history.keys()):
            # add center name and week index columns
            centerwise_df['center'] = c_name
            centerwise_df['week'] = range(1, FoodDelivery.__num_weeks + 1)

        df_per_center_statistics = pd.concat(arr_df_per_center_statistics, axis=0)
        
        if self.__input_mapper is not None:
            df_per_center_statistics['center'] = df_per_center_statistics['center'].map(
                {v: k for k, v in self.__input_mapper.items()}
            )

        simRes = FoodDeliveryResult(score, df_aggregated_statistics, df_per_center_statistics)
        return simRes
